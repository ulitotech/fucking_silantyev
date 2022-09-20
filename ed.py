"""Программа для автоматического формирования документации для Силантьева"""

# Импортируем необходимые библиотеки

import random
from random import uniform
import openpyxl as xl
import docx2pdf as d2p
import os
import docx
from docxtpl import DocxTemplate
from PyPDF2 import PdfMerger
import xlwings as xw
import datetime
from docxtpl import InlineImage

# Прописываем константы и пути

NOT_BALANCE = 20
PATTERN_NAMES = ('00_ТЛ, Обложка.docx', '02_ОУ.docx', '03_ РСО.docx', '04_АО.docx', '05_АСП.docx',
                 '06_АВ.docx', '06_АЭ.docx', '08_РМ.docx', '09_ВОП.docx', '10_ОЧ.docx', '11_П.docx', '12_ПС.docx',
                 '13_Ж.docx', '01_ВД.docx')
XCL_NAMES = ('all_data.xlsx', 'ks_2.xlsx')
IMAGE_NAMES = ("001.jpg", "002.jpg")
PATTERN_PATH = r'C:\Users\ulito\Desktop\training\patterns\ex_d\\'
PATH_OUT = r"C:\Users\ulito\Desktop\training\out\ex\\"
PATH_XCL = r"C:\Users\ulito\Desktop\training\xl\\"
PATH_PASS = r"C:\Users\ulito\Desktop\training\patterns\ex_d\passports\\"
PATH_IMAGE = r"C:\Users\ulito\Desktop\training\patterns\ex_d\images\\"
size_doc = [2,2,2,2]

# Функции

def file_convert_docx_pdf(dirs):
    """Конвертируем файл в PDF"""

    #Список файлов в директории
    file_in_dir = os.listdir(dirs)

    #Конвертация...
    for file in file_in_dir:
        if file.endswith('.docx'):
            file_k = f'{file.split(".")[0].replace(".", "_")}.pdf'
            d2p.convert(f'{dirs}\\{file}', f'{dirs}\\{file_k}')
        else:
            continue

def pdf_merge(path_in):
    """Объединяем все PDF в один"""
    # Список файлов в директории
    content = os.listdir(path_in)
    pdfs = []
    # Отбор pdf
    for file in content:
        if os.path.isfile(os.path.join(path_in, file)) and file.endswith('.pdf'):
            pdfs.append(file)
    merger = PdfMerger()
    #Объединение
    for pdf in pdfs:
        merger.append(f"{path_in}{pdf}")
    merger.write(f"{path_in}СЭК.421736.03.46.{tp}.ИД.pdf")
    merger.close()

# Заполняем тэги в документах и сохраняем в итоговой папке

tp = 257
wb = xl.load_workbook(f"{PATH_XCL}\{XCL_NAMES[0]}", read_only=True)
sheets = wb.sheetnames
ws=wb[sheets[1]]
tp_info = []
for row in ws.iter_rows():
    for cell in row:
        if cell.value == tp:
            for cell in row:
                if cell.value == None:
                    tp_info.append("")
                else:
                    tp_info.append(cell.value)
ws=wb[sheets[0]]
inst_info = []
for row in ws.iter_rows():
    for cell in row:
        if cell.value == tp:
            inst_info.append([c.value if c.value != None else "" for c in row])
ws=wb[sheets[2]]
uspd = []
for row in ws.iter_rows():
    for cell in row:
        if cell.value == tp:
            uspd.append([c.value for c in row])
inst_date = []
for row in inst_info[1:]:
        inst_date.append(datetime.datetime.strptime(row[16],"%d.%m.%Y").date())
beg_inst = min(inst_date).strftime("%d.%m.%Y")
end_inst = max(inst_date).strftime("%d.%m.%Y")
ttr_1_ib = ttr_1_iz = ttr_1_nb = ttr_1_nz = ttr_3_ib = ttr_3_iz = ttr_3_nb = ttr_3_nz = ttr_1_ib_i = ttr_1_iz_i =\
    ttr_1_nb_i = ttr_1_nz_i = ttr_3_ib_i = ttr_3_iz_i = ttr_3_nb_i = ttr_3_nz_i = 0

# Подсчитываем количество ТТР

for row in inst_info:
    if row[11] == "3/иб":
        if "2400/5" in row[12]:
            ttr_3_ib_i += 1
        else:
            ttr_3_ib += 1
    if row[11] == "3/из":
        if "2400/5" in row[12]:
            ttr_3_iz_i += 1
        else:
            ttr_3_iz += 1
    if row[11] == "3/нб":
        if "2400/5" in row[12]:
            ttr_3_nb_i += 1
        else:
            ttr_3_nb += 1
    if row[11] == "3/нз":
        if "2400/5" in row[12]:
            ttr_3_nz_i += 1
        else:
            ttr_3_nz += 1
    if row[11] == "1/иб":
        if "2400/5" in row[12]:
            ttr_1_ib_i += 1
        else:
            ttr_1_ib += 1
    if row[11] == "1/из":
        if "2400/5" in row[12]:
            ttr_1_iz_i += 1
        else:
            ttr_1_iz += 1
    if row[11] == "1/нб":
        if "2400/5" in row[12]:
            ttr_1_nb_i += 1
        else:
            ttr_1_nb += 1
    if row[11] == "1/нз":
        if "2400/5" in row[12]:
            ttr_1_nz_i += 1
        else:
            ttr_1_nz += 1

# Изменяем количество ТТР для расчета

xl_b = xw.Book(f"{PATH_XCL}\{XCL_NAMES[1]}")
wks = xw.sheets["Source"]
wks.range('U9').value = ttr_1_ib_i
wks.range('V9').value = ttr_1_iz_i
wks.range('W9').value = ttr_1_nb_i
wks.range('X9').value = ttr_1_nz_i
wks.range('Y9').value = ttr_3_ib_i
wks.range('Z9').value = ttr_3_iz_i
wks.range('AA9').value = ttr_3_nb_i
wks.range('AB9').value = ttr_3_nz_i
wks.range('U16').value = ttr_1_ib
wks.range('V16').value = ttr_1_iz
wks.range('W16').value = ttr_1_nb
wks.range('X16').value = ttr_1_nz
wks.range('Y16').value = ttr_3_ib
wks.range('Z16').value = ttr_3_iz
wks.range('AA16').value = ttr_3_nb
wks.range('AB16').value = ttr_3_nz
xl_b.save()
xl_b.close()

# Генерим таблицу 8

wb = xl.load_workbook(f"{PATH_XCL}\{XCL_NAMES[1]}", data_only=True)
sheets = wb.sheetnames
ws=wb[sheets[1]]
speka = []
for i in range(1,20):
    if ws.cell(row = i, column = 4).value != 0:
        speka.append([ws.cell(row = i, column = j).value for j in range(2,5)])
table_08 = []
for i in range(len(speka)):
    table_08.append({
        'Index': i+1,
        'Name': speka[i][0],
        'Measure': speka[i][1],
        'Quantity': speka[i][2],
    })

# Генерим таблицу 3

table_03 = []
for i in range(len(inst_info)):
    table_03.append({
        'Index': i+1,
        'Res': "Истринский РЭС",
        'Tp': f"ТП-{inst_info[i][1]}",
        'City': inst_info[i][4],
        'Street': inst_info[i][5],
        'Home': inst_info[i][6],
        'Part': "",
        'Build': "",
        'Flat': inst_info[i][7],
        'Name': inst_info[i][9],
        'TypeU': inst_info[i][3],
        'Post': inst_info[i][2],
        'Gps': inst_info[i][10],
        'Ttr': inst_info[i][11],
        'TypePwr': inst_info[i][12],
        'Serial': inst_info[i][13]
        })
table_03.append({
    'Index': len(table_03)+1,
    'Res': "Истринский РЭС",
    'Tp': f"ТП-{inst_info[0][1]}",
    'City': inst_info[0][4],
    'Street': "",
    'Home': "",
    'Part': "",
    'Build': "",
    'Flat': "",
    'Name': "",
    'TypeU': "",
    'Post': "",
    'Gps': "",
    'Ttr': "",
    'TypePwr': uspd[0][1],
    'Serial': f"{uspd[0][2]}, IP={tp_info[4]}"
    })
table_03.append({
    'Index': len(table_03)+1,
    'Res': "Истринский РЭС",
    'Tp': f"ТП-{inst_info[0][1]}",
    'City': inst_info[0][4],
    'Street': "",
    'Home': "",
    'Part': "",
    'Build': "",
    'Flat': "",
    'Name': "",
    'TypeU': "",
    'Post': "",
    'Gps': "",
    'Ttr': "",
    'TypePwr': uspd[0][3],
    'Serial': f"{uspd[0][4]}"
    })

# Генерим таблицу 12

passport = []
table_12 = []
wb = xl.load_workbook(f"{PATH_PASS}p_{tp}.xlsx", data_only=True)
sheets = wb.sheetnames
ws=wb[sheets[0]]
for row in ws.iter_rows():
        passport.append([c.value if c.value != None else "" for c in row])
for i in range(1,len(passport)):
    table_12.append({'Address': passport[i][0],
                     'Type': passport[i][1],
                     'Name': passport[i][2],
                     'Income': passport[i][3],
                    'Asdu' : passport[i][4]
    })

# Генерируем таблицу 13

journal = []
table_13 = []
real_pass_date = (datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()) \
    if datetime.datetime.weekday(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y"))+1 not in [6,7]\
    else (datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=2))\
    if datetime.datetime.weekday(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y"))+1==6\
    else (datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=1))
for i in range(31):
    journal_row = []
    journal_row.append(i)
    journal_row.append((real_pass_date+datetime.timedelta(days=i)).strftime("%d.%m.%Y"))
    b = round(random.uniform(100-NOT_BALANCE-2,100-NOT_BALANCE+2),2)
    if b <= 100:
        journal_row.append(b)
    else:
        journal_row.append(100)
    journal_row.append(round(random.uniform(95, 100),2))
    journal.append(journal_row)
for i in range(30):
    table_13.append({'Index': journal[i][0]+1,
                     'Date': journal[i][1],
                     'Balance': f"{journal[i][2]}%",
                     'Survey': f"{journal[i][3]}%",
                     'Result': "Условия выполнены" if (journal[i][2]>=88 and journal[i][3]>=95)
                     else "Условия не выполнены",
                     'Name' : "Силантьев В.В."
                     })
res_1 = f"Обобщенные выводы по результатам проведенной опытной эксплуатации: " \
        f"\n1. Величина опроса У не менее 95% \n2. Величина баланса распределения ЭЭ не объекте {100-NOT_BALANCE}%"
res_2 = "\tНебаланс более 12% по объекту обусловлен наличием ПУ с GSM-модемом," \
                                                   " установленных в рамках исполнения договоров М/3634-СЭК от 20.11.2020," \
                                                   " М/3635-СЭК от 20.11.2020, РМР/2097-СЭК 03.06.2021, РМР/2351-СЭК 15.06.2021," \
                                                   " РМР/2109-СЭК 03.06.2021, РМР/2424-СЭК 17.06.2021, РМР/2425-СЭК 17.06.2021,РМР/2423-СЭК 17.06.2021," \
                                                   " РМР 1575-СЭК от 11.07.2022, 21/4/28-1 от 28.04.2021, 22/8/22-1 от 22.08.2022. ," \
                                                   " проведение пуско-наладочных работ по которым не входят в согласованные договором объемы работ по проекту."
sum_result = res_1 if NOT_BALANCE < 12 else res_2
counter = 0
for i in PATTERN_NAMES:
    # Создание директории ТП (при отсутствии)
    if not os.path.isdir(f'{PATH_OUT}\ТП-{tp}'):
        os.mkdir(f'{PATH_OUT}\ТП-{tp}')
    doc = DocxTemplate(f"{PATTERN_PATH}{i}")
    n_image = 0 if "TOPAZ" in uspd[0][1] else 1
    image = InlineImage(doc, f"{PATH_IMAGE}{IMAGE_NAMES[n_image]}")
    context = {"n_tp" : tp_info[2],
               "city" : tp_info[7],
               "pass_date" : f'{(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=-1)).strftime("%d.%m.%Y") if datetime.datetime.weekday(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y"))+1 not in [6,7] else (datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=-2)).strftime("%d.%m.%Y")}',
               "real_pass_date" : f'{(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()).strftime("%d.%m.%Y") if datetime.datetime.weekday(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y"))+1 not in [6,7] else (datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=2)).strftime("%d.%m.%Y") if datetime.datetime.weekday(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y"))+1==6 else (datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=1)).strftime("%d.%m.%Y")}',
               "beg_inst" : beg_inst,
               "end_inst" : end_inst,
               "prom_date" : f'{(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=30)).strftime("%d.%m.%Y") if datetime.datetime.weekday(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y"))+1 not in [6,7] else (datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=32)).strftime("%d.%m.%Y") if datetime.datetime.weekday(datetime.datetime.strptime(tp_info[12],"%d.%m.%Y"))+1==6 else (datetime.datetime.strptime(tp_info[12],"%d.%m.%Y").date()+datetime.timedelta(days=31)).strftime("%d.%m.%Y")}',
               "image_place" : image,
               "table_03" : table_03,
               "table_08" : table_08,
               "table_12" : table_12,
               "table_13" : table_13,
               "sum_result" : sum_result,
               "p_3": size_doc[0],
               "p_8": size_doc[1],
               "p_12": size_doc[2],
               "p_13": size_doc[3]
               }
    doc.render(context)
    doc.save(f"{PATH_OUT}ТП-{tp}\{i}")
print("done_doc")

# Конвертируем в PDF

file_convert_docx_pdf(f"{PATH_OUT}ТП-{tp}\\")

# Объединяем в один PDF файл

pdf_merge(f"{PATH_OUT}\ТП-{tp}\\")
print("done_merge")