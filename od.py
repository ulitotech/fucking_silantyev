"""Программа для автоматического формирования документации для Силантьева"""

# Импортируем необходимые библиотеки
import random

import openpyxl as xl
import docx2pdf as d2p
import os
import docx
from docxtpl import DocxTemplate
from PyPDF2 import PdfMerger
import xlwings as xw
import datetime
from docxtpl import InlineImage

# Прописываем константы и пути

PATTERN_NAMES = ('00_ТЛ, Обложка.docx', '02_РП.docx', '03_ Р.docx', '04_РЭ.docx', '01_ВД.docx')
XCL_NAMES = 'all_data.xlsx'
PATTERN_PATH = r'C:\Users\ulito\Desktop\training\patterns\op_d\\'
PATH_OUT = r"C:\Users\ulito\Desktop\training\out\op\\"
PATH_XCL = r"C:\Users\ulito\Desktop\training\xl\\"
PATH_PASS = r"C:\Users\ulito\Desktop\training\patterns\ex_d\passports\\"

# Функции

def file_convert_docx_pdf(dirs):
    """Конвертируем файл в PDF"""

    #Список файлов в директории
    file_in_dir = os.listdir(dirs)

    #Конвертация...
    for file in file_in_dir:
        if file.endswith('.docx'):
            file_k = f'{file.split(".")[0].replace(".", "_")}.pdf'
            d2p.convert(f'{dirs}\\{file}', f'{dirs}\\{file_k}')
        else:
            continue

def pdf_merge(path_in):
    """Объединяем все PDF в один"""
    # Список файлов в директории
    content = os.listdir(path_in)
    pdfs = []
    # Отбор pdf
    for file in content:
        if os.path.isfile(os.path.join(path_in, file)) and file.endswith('.pdf'):
            pdfs.append(file)
    merger = PdfMerger()
    #Объединение
    for pdf in pdfs:
        merger.append(f"{path_in}{pdf}")
    merger.write(f"{path_in}СЭК.421736.03.46.{tp}.ЭД.pdf")
    merger.close()

# Заполняем тэги в документах и сохраняем в итоговой папке

TP = [2262]
for tp in TP:
    wb = xl.load_workbook(f"{PATH_XCL}{XCL_NAMES}", read_only=True)
    sheets = wb.sheetnames
    ws=wb[sheets[1]]
    tp_info = []
    tp_list = []
    for row in ws.iter_rows():
        tp_list.append(row[2].value)
    tp_list = tp_list[1:]
    for row in ws.iter_rows():
        for cell in row[:3]:
            if cell.value == tp:
                for cell in row:
                    if cell.value == None:
                        tp_info.append("")
                    else:
                        tp_info.append(cell.value)

    ws=wb[sheets[0]]
    inst_info = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == tp:
                inst_info.append([c.value if c.value != None else "" for c in row])
    ws=wb[sheets[2]]
    uspd = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == tp:
                uspd.append([c.value for c in row])
    inst_date = []
    for row in inst_info[:]:
            inst_date.append(datetime.datetime.strptime(row[16],"%d.%m.%Y").date())

    end_inst = max(inst_date).strftime("%d.%m.%Y") if datetime.datetime.weekday(max(inst_date))+1 not in [6,7]  \
        else (max(inst_date)+datetime.timedelta(days=2)).strftime("%d.%m.%Y") if datetime.datetime.weekday(max(inst_date))+1 == 6 \
        else (max(inst_date)+datetime.timedelta(days=1)).strftime("%d.%m.%Y")

    # Генерим таблицу 2

    table_02 = []
    for i in range(len(inst_info)):
        table_02.append({
            'Index': i+1,
            'Tp': f"ТП-{inst_info[i][1]}",
            'Serial': inst_info[i][13],
            'TypePwr': inst_info[i][12],
            'TypeU': inst_info[i][3]
            })
    table_02.append({
        'Index': len(table_02)+1,
        'Tp': f"ТП-{inst_info[i][1]}",
        'Serial': f"{uspd[0][2]}, IP={tp_info[4]}",
        'TypePwr': uspd[0][1],
        'TypeU': "-"
        })
    table_02.append({
        'Index': len(table_02)+1,
        'Tp': f"ТП-{inst_info[i][1]}",
        'Serial': f"{uspd[0][4]}",
        'TypePwr': uspd[0][3],
        'TypeU': "-"
        })

    # Генерим таблицу 3

    table_03 = []
    for i in range(len(inst_info)):
        table_03.append({
            'Index': i+1,
            'Tp': f"ТП-{inst_info[i][1]}",
            'Res': "Истринский РЭС",
            'City': inst_info[i][4],
            'Get_EU' : end_inst,
            'Name': inst_info[i][9],
            'Street': inst_info[i][5],
            'Home': inst_info[i][6],
            'Flat': inst_info[i][7],
            'Phone': "",
            'Ls' : inst_info[i][8],
            'TypePwr': inst_info[i][12],
            'Serial': inst_info[i][13],
            'Plomb' : inst_info[i][15],
            'TypePwrOld' : inst_info[i][19],
            'SerialOld' : inst_info[i][18]
            })

    # Заполняем шаблоны

    for i in PATTERN_NAMES:
        # Создание директории ТП (при отсутствии)
        if not os.path.isdir(f'{PATH_OUT}\ТП-{tp}'):
            os.mkdir(f'{PATH_OUT}\ТП-{tp}')
        doc = DocxTemplate(f"{PATTERN_PATH}{i}")
        context = {"n_tp" : tp_info[2],
                   "city" : tp_info[7],
                   "table_03" : table_03,
                   "table_02" : table_02
                   }
        doc.render(context)
        doc.save(f"{PATH_OUT}ТП-{tp}\{i}")

    print(f"ТП-{tp}_done_doc")

    # Конвертируем в PDF

    file_convert_docx_pdf(f"{PATH_OUT}ТП-{tp}\\")

    # Объединяем в один PDF файл

    pdf_merge(f"{PATH_OUT}\ТП-{tp}\\")
    print(f"ТП-{tp}_done_merge")
print(f"Эксплутационная документация по всем ТП сделана")
